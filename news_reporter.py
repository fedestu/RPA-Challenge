import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import logging
import os
from datetime import datetime

class NewsReporter:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    def create_excel_report(self, data):
        """Create an Excel report from the given data."""
        # Get today's date in 'YYYY-MM-DD' format
        today_str = datetime.now().strftime('%Y-%m-%d')

        # File name including today's date
        filename = os.path.join(self.output_dir, f"news_data_{today_str}.xlsx")

        # Ensure the output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

        logging.info("Creating Excel report")

        # Create a DataFrame and save it to an Excel file
        df = pd.DataFrame(data)
        df.to_excel(filename, index=False, engine='openpyxl')

        # Load the workbook to apply formatting
        wb = load_workbook(filename)
        ws = wb.active

        # Bold column names
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logging.warning(f"Error adjusting column width: {e}")
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filename)

        logging.info(f"Excel report created successfully at {filename}")
